from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from order.models import Order, OrderItem
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDate
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import io
import logging
logger = logging.getLogger('salesreport.views')


def is_staff(user):
    return user.is_staff


@login_required(login_url='accounts:admin_login')
@never_cache
@user_passes_test(is_staff, 'accounts:admin_login')
def view_salesreport(request):
    logger.debug("view_salesreport called | user=%s | method=%s", request.user, request.method)

    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    sort_option = request.GET.get('sortOption')

    # Parse dates
    if startdate:
        try:
            startdate = datetime.strptime(startdate, '%Y-%m-%d').date()
            logger.debug("Parsed startdate: %s", startdate)
        except ValueError:
            logger.warning("Invalid startdate format: %s", startdate)
            startdate = None
    if enddate:
        try:
            enddate = datetime.strptime(enddate, '%Y-%m-%d').date()
            logger.debug("Parsed enddate: %s", enddate)
        except ValueError:
            logger.warning("Invalid enddate format: %s", enddate)
            enddate = None

    # Base query
    sales_report = Order.objects.annotate(date=TruncDate('created_at')).values('date').annotate(
        total_sales_revenue=Sum('total_price'),
        coupon_applied=Sum('coupon_amount'),
        offer_applied=Sum('total_offer_discount'),
        net_sales=Sum('total_price') - Sum(F('coupon_amount') + F('total_offer_discount')),
        number_of_orders=Count('order_id'),
        total_items_sold=Sum('items__quantity', default=0)
    ).order_by('date')

    # Apply date filters
    if startdate and enddate:
        sales_report = sales_report.filter(date__range=[startdate, enddate])
        logger.info("Filtered by date range: %s to %s", startdate, enddate)
    elif startdate:
        sales_report = sales_report.filter(date__gte=startdate)
        logger.info("Filtered from startdate: %s", startdate)
    elif enddate:
        sales_report = sales_report.filter(date__lte=enddate)
        logger.info("Filtered up to enddate: %s", enddate)

    # Apply sort option
    today = date.today()
    if sort_option == 'day':
        sales_report = sales_report.filter(date=today)
        logger.info("Filtered: Today (%s)", today)
    elif sort_option == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        sales_report = sales_report.filter(date__gte=start_of_week)
        logger.info("Filtered: This week starting %s", start_of_week)
    elif sort_option == 'month':
        sales_report = sales_report.filter(date__month=today.month, date__year=today.year)
        logger.info("Filtered: Current month (%s-%s)", today.year, today.month)
    elif sort_option == 'year':
        sales_report = sales_report.filter(date__year=today.year)
        logger.info("Filtered: Current year (%s)", today.year)

    # Aggregated totals
    total_sales_price = OrderItem.objects.exclude(status='Cancelled').aggregate(Sum('price'))['price__sum'] or 0
    total_order = Order.objects.count()
    total_unit_sold = int(OrderItem.objects.exclude(status='Cancelled').aggregate(Sum('quantity'))['quantity__sum'] or 0)
    total_order_amount = Order.objects.aggregate(Sum('total_price'))['total_price__sum'] or 0
    total_discount_amount = Order.objects.annotate(
        total_discount=F('coupon_amount') + F('total_offer_discount')
    ).aggregate(total_discount_amount=Sum('total_discount'))['total_discount_amount'] or 0

    logger.debug("Aggregated totals | orders=%d | units=%d | revenue=%s", total_order, total_unit_sold, total_order_amount)

    context = {
        'total_sales_price': total_sales_price,
        'total_order': total_order,
        'total_unit_sold': total_unit_sold,
        'total_order_amount': total_order_amount,
        'total_discount_amount': total_discount_amount,
        'sales_report': sales_report,
        'startdate': startdate,
        'enddate': enddate,
        'sortOption': sort_option,
    }

    # Handle download requests
    if 'download_pdf' in request.GET:
        logger.info("PDF download requested | user=%s", request.user)
        return download_pdf(sales_report, startdate, enddate, sort_option)
    elif 'download_excel' in request.GET:
        logger.info("Excel download requested | user=%s", request.user)
        return download_excel(sales_report)

    logger.info("Sales report page rendered | user=%s | records=%d", request.user, len(list(sales_report)))
    return render(request, 'admin/salesreport.html', context)


def download_pdf(sales_report, startdate, enddate, sort_option):
    logger.debug("download_pdf started | records=%d", len(list(sales_report)))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'Header', fontSize=10, textColor=colors.black, spaceAfter=6, alignment=1
    )
    title_style = ParagraphStyle('Title', fontSize=14, spaceAfter=12, alignment=1)

    company_info = Paragraph(
        "<b>BRAT Mens Clothing PVT</b><br/>"
        "Email: bratclothing@gmail.com<br/>"
        "Website: www.bratmenswear.com", header_style
    )
    elements.append(company_info)
    elements.append(Spacer(1, 12))

    title = f"Sales Report ({startdate or 'All'} - {enddate or 'All'}) - Sorted by {sort_option.capitalize() if sort_option else 'Date'}"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Recalculate totals for PDF
    total_sales_price = OrderItem.objects.exclude(status='Cancelled').aggregate(Sum('price'))['price__sum'] or 0
    total_order = Order.objects.count()
    total_unit_sold = int(OrderItem.objects.exclude(status='Cancelled').aggregate(Sum('quantity'))['quantity__sum'] or 0)
    total_order_amount = Order.objects.aggregate(Sum('total_price'))['total_price__sum'] or 0
    total_discount_amount = Order.objects.annotate(
        total_discount=F('coupon_amount') + F('total_offer_discount')
    ).aggregate(total_discount_amount=Sum('total_discount'))['total_discount_amount'] or 0

    summary_data = [
        ["Total Sales Price", f"{total_sales_price:.2f}"],
        ["Total Orders", total_order],
        ["Total Units Sold", total_unit_sold],
        ["Total Order Amount", f"{total_order_amount:.2f}"],
        ["Total Discount Amount", f"{total_discount_amount:.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    table_data = [
        ["Date", "Total Sales Revenue", "Coupon Applied", "Offer Applied",
         "Net Sales", "Number of Orders", "Total Items Sold"]
    ]

    for record in sales_report:
        table_data.append([
            record['date'].strftime('%Y-%m-%d') if record['date'] else 'N/A',
            f"{record['total_sales_revenue'] or 0:.2f}",
            f"{record['coupon_applied'] or 0:.2f}",
            f"{record['offer_applied'] or 0:.2f}",
            f"{record['net_sales'] or 0:.2f}",
            record['number_of_orders'],
            record['total_items_sold']
        ])

    table = Table(table_data, colWidths=[1*inch] + [1.1*inch]*6)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    try:
        doc.build(elements)
        buffer.seek(0)
        logger.info("PDF generated successfully | size=%d bytes", buffer.getbuffer().nbytes)
    except Exception as e:
        logger.error("PDF generation failed | error=%s", e, exc_info=True)
        return HttpResponse("Error generating PDF", status=500)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    return response


def download_excel(sales_report):
    logger.debug("download_excel started | records=%d", len(list(sales_report)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws['A1'] = "Summary"
    ws['A1'].font = Font(bold=True)

    # Recalculate totals
    total_sales_price = OrderItem.objects.exclude(status='Cancelled').aggregate(Sum('price'))['price__sum'] or 0
    total_order = Order.objects.count()
    total_unit_sold = int(OrderItem.objects.exclude(status='Cancelled').aggregate(Sum('quantity'))['quantity__sum'] or 0)
    total_order_amount = Order.objects.aggregate(Sum('total_price'))['total_price__sum'] or 0
    total_discount_amount = Order.objects.annotate(
        total_discount=F('coupon_amount') + F('total_offer_discount')
    ).aggregate(total_discount_amount=Sum('total_discount'))['total_discount_amount'] or 0

    summary = [
        ("Total Sales Price", total_sales_price),
        ("Total Orders", total_order),
        ("Total Units Sold", total_unit_sold),
        ("Total Order Amount", total_order_amount),
        ("Total Discount Amount", total_discount_amount)
    ]
    for row_num, (label, value) in enumerate(summary, start=2):
        ws[f'A{row_num}'] = label
        ws[f'B{row_num}'] = value
        ws[f'A{row_num}'].font = Font(bold=True)

    headers = ["Date", "Revenue", "Coupon Applied", "Offer Applied", "Net Sales", "Orders Count", "Items Sold"]
    column_widths = [20, 20, 18, 18, 18, 15, 15]
    ws.append(headers)
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    for record in sales_report:
        ws.append([
            record['date'],
            record['total_sales_revenue'] or 0,
            record['coupon_applied'] or 0,
            record['offer_applied'] or 0,
            record['net_sales'] or 0,
            record['number_of_orders'],
            record['total_items_sold']
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    try:
        wb.save(response)
        logger.info("Excel generated successfully")
    except Exception as e:
        logger.error("Excel generation failed | error=%s", e, exc_info=True)
        return HttpResponse("Error generating Excel", status=500)

    return response